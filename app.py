"""
Abiball 2026 – Reservation Backend
Flask + SQLite + ReportLab PDF + SMTP email
"""

import os, io, json, sqlite3, smtplib, threading, random, string
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from flask import Flask, request, jsonify, send_file, send_from_directory, g
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.utils import ImageReader
import qrcode
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

# ── Config ───────────────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
app = Flask(__name__, static_folder=os.path.join(BASE_DIR, 'static'))
DB_PATH    = os.path.join(BASE_DIR, 'db', 'abiball.db')
SMTP_HOST  = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
SMTP_PORT  = int(os.environ.get('SMTP_PORT', '587'))
SMTP_USER  = os.environ.get('SMTP_USER', '')
SMTP_PASS  = os.environ.get('SMTP_PASS', '')
ADMIN_PW   = os.environ.get('ADMIN_PW',  'AbiVegas2026admin')
EMAIL_ENABLED = bool(SMTP_USER and SMTP_PASS)

TABLES = [
    {"id":  1, "name": "Lehrertisch", "icon": "A1", "max_seats": 4, "teacher": True},
    {"id":  2, "name": "Tisch A2", "icon": "A2", "max_seats": 4},
    {"id":  3, "name": "Tisch A3", "icon": "A3", "max_seats": 4},
    {"id":  4, "name": "Tisch A4", "icon": "A4", "max_seats": 4},
    {"id":  5, "name": "Tisch A5", "icon": "A5", "max_seats": 4},
    {"id":  6, "name": "Tisch B1", "icon": "B1", "max_seats": 4},
    {"id":  7, "name": "Tisch B2", "icon": "B2", "max_seats": 4},
    {"id":  8, "name": "Tisch B3", "icon": "B3", "max_seats": 4},
    {"id":  9, "name": "Tisch B4", "icon": "B4", "max_seats": 4},
    {"id": 10, "name": "Tisch B5", "icon": "B5", "max_seats": 4},
    {"id": 11, "name": "Tisch C1", "icon": "C1", "max_seats": 4},
    {"id": 12, "name": "Tisch C2", "icon": "C2", "max_seats": 4},
    {"id": 13, "name": "Tisch C3", "icon": "C3", "max_seats": 4},
    {"id": 14, "name": "Tisch C4", "icon": "C4", "max_seats": 4},
    {"id": 15, "name": "Tisch C5", "icon": "C5", "max_seats": 4},
    {"id": 16, "name": "Tisch D1", "icon": "D1", "max_seats": 4},
    {"id": 17, "name": "Tisch D2", "icon": "D2", "max_seats": 4},
    {"id": 18, "name": "Tisch D3", "icon": "D3", "max_seats": 4},
    {"id": 19, "name": "Tisch D4", "icon": "D4", "max_seats": 4},
    {"id": 20, "name": "Tisch D5", "icon": "D5", "max_seats": 4},
    {"id": 21, "name": "Tisch E1", "icon": "E1", "max_seats": 4},
    {"id": 22, "name": "Tisch E2", "icon": "E2", "max_seats": 4},
    {"id": 23, "name": "Tisch E3", "icon": "E3", "max_seats": 4},
    {"id": 24, "name": "Tisch E4", "icon": "E4", "max_seats": 4},
    {"id": 25, "name": "Tisch E5", "icon": "E5", "max_seats": 4},
    {"id": 26, "name": "Tisch C6", "icon": "C6", "max_seats": 4},
    {"id": 27, "name": "Tisch C7", "icon": "C7", "max_seats": 4},
    {"id": 28, "name": "Tisch D6", "icon": "D6", "max_seats": 4},
    {"id": 29, "name": "Tisch D7", "icon": "D7", "max_seats": 4},
    {"id": 30, "name": "Tisch E6", "icon": "E6", "max_seats": 4},
    {"id": 31, "name": "Tisch E7", "icon": "E7", "max_seats": 4},
    {"id": 32, "name": "Tisch C8", "icon": "C8", "max_seats": 4},
    {"id": 33, "name": "Tisch C9", "icon": "C9", "max_seats": 4},
    {"id": 34, "name": "Tisch D8", "icon": "D8", "max_seats": 4},
    {"id": 35, "name": "Tisch D9", "icon": "D9", "max_seats": 4},
    {"id": 36, "name": "Tisch E8", "icon": "E8", "max_seats": 4},
    {"id": 37, "name": "Tisch E9", "icon": "E9", "max_seats": 4},
    {"id": 38, "name": "Tisch A10", "icon": "A10", "max_seats": 4},
    {"id": 39, "name": "Tisch A11", "icon": "A11", "max_seats": 4},
    {"id": 40, "name": "Tisch A12", "icon": "A12", "max_seats": 4},
    {"id": 41, "name": "Tisch A13", "icon": "A13", "max_seats": 4},
    {"id": 42, "name": "Tisch B10", "icon": "B10", "max_seats": 4},
    {"id": 43, "name": "Tisch B12", "icon": "B12", "max_seats": 4},
    {"id": 44, "name": "Tisch B13", "icon": "B13", "max_seats": 4},
    {"id": 45, "name": "Tisch B14", "icon": "B14", "max_seats": 4},
    {"id": 46, "name": "Tisch C10", "icon": "C10", "max_seats": 4},
    {"id": 47, "name": "Tisch C11", "icon": "C11", "max_seats": 4},
    {"id": 48, "name": "Tisch C12", "icon": "C12", "max_seats": 4},
    {"id": 49, "name": "Tisch C13", "icon": "C13", "max_seats": 4},
    {"id": 50, "name": "Tisch M1", "icon": "D10", "max_seats": 4},
    {"id": 51, "name": "Tisch M2", "icon": "D11", "max_seats": 4},
    {"id": 52, "name": "Tisch M3", "icon": "D12", "max_seats": 4},
    {"id": 53, "name": "Tisch M4", "icon": "D13", "max_seats": 4},
    {"id": 54, "name": "Tisch N1", "icon": "E10", "max_seats": 4},
    {"id": 55, "name": "Tisch N2", "icon": "E11", "max_seats": 4},
    {"id": 56, "name": "Tisch N3", "icon": "E12", "max_seats": 4},
    {"id": 57, "name": "Tisch N4", "icon": "E13", "max_seats": 4},
    {"id": 58, "name": "Tisch F1", "icon": "F1", "max_seats": 4},
    {"id": 59, "name": "Tisch F2", "icon": "F2", "max_seats": 4},
    {"id": 60, "name": "Tisch F3", "icon": "F3", "max_seats": 4},
    {"id": 61, "name": "Tisch F4", "icon": "F4", "max_seats": 4},
    {"id": 62, "name": "Tisch F5", "icon": "F5", "max_seats": 4},
    {"id": 63, "name": "Tisch G1", "icon": "G1", "max_seats": 4},
    {"id": 64, "name": "Tisch G2", "icon": "G2", "max_seats": 4},
    {"id": 65, "name": "Tisch G3", "icon": "G3", "max_seats": 4},
    {"id": 66, "name": "Tisch G4", "icon": "G4", "max_seats": 4},
    {"id": 67, "name": "Tisch G5", "icon": "G5", "max_seats": 4},
    {"id": 68, "name": "Tisch F6", "icon": "F6", "max_seats": 4},
    {"id": 69, "name": "Tisch F7", "icon": "F7", "max_seats": 4},
    {"id": 70, "name": "Tisch G6", "icon": "G6", "max_seats": 4},
    {"id": 71, "name": "Tisch G7", "icon": "G7", "max_seats": 4},
    {"id": 72, "name": "Tisch F8", "icon": "F8", "max_seats": 4},
    {"id": 73, "name": "Tisch F9", "icon": "F9", "max_seats": 4},
    {"id": 74, "name": "Tisch G8", "icon": "G8", "max_seats": 4},
    {"id": 75, "name": "Tisch G9", "icon": "G9", "max_seats": 4},
    {"id": 76, "name": "Tisch F11", "icon": "F11", "max_seats": 4},
    {"id": 77, "name": "Tisch F12", "icon": "F12", "max_seats": 4},
    {"id": 78,("id"): 78,("name"): ("Tisch F13"),("icon"): ("F13"),("max_seats"): (4)},
    {"id": 79, "name": "Tisch F14",("icon"): ("F14"),("max_seats"): (4)},
    {"id": 80,("name"): ("Tisch G11"),("icon"): ("G11"),("max_seats"): (4),("id"): (80)},
    {"id": 81, "name": "Tisch G12", "icon": "G12", "max_seats": 4},
    {"id": 82, "name": "Tisch G13", "icon": "G13", "max_seats": 4},
    {"id": 83, "name": "Tisch G14", "icon": "G14", "max_seats": 4},
    {"id": 84, "name": "Tisch F10", "icon": "F10", "max_seats": 4},
    {"id": 85, "name": "Tisch G10", "icon": "G10", "max_seats": 4},
]

# ── How to add more tables ────────────────────────────────────────────────────
# To add more tables, append new entries to the TABLES list above.
# Each table is a dictionary with the following keys:
#   - "id": A unique integer ID (increment from the last one, e.g., 58, 59, etc.)
#   - "name": A descriptive name for the table (e.g., "Tisch N1")
#   - "icon": A short icon/label (e.g., "N1")
#   - "max_seats": The maximum number of seats at the table (integer)
#
# Example of adding a new table:
# TABLES.append({"id": 58, "name": "Tisch N1", "icon": "N1", "max_seats": 8})
#
# Note: If adding tables beyond the current layout (IDs 1-57), you may need to update
# the tablePos() function in static/index.html to position the new tables correctly
# on the floor plan SVG.

# ── DB ────────────────────────────────────────────────────────────────────────
def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
    return g.db

@app.teardown_appcontext
def close_db(exc):
    db = g.pop('db', None)
    if db: db.close()

def init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    con = sqlite3.connect(DB_PATH)
    con.execute("""
        CREATE TABLE IF NOT EXISTS reservations (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            table_id    INTEGER NOT NULL,
            table_name  TEXT    NOT NULL,
            name        TEXT    NOT NULL,
            email       TEXT    NOT NULL,
            seats       INTEGER NOT NULL,
            created_at  TEXT    NOT NULL,
            conf_code   TEXT    NOT NULL UNIQUE
        )
    """)
    con.commit(); con.close()

# ── PDF ────────────────────────────────────────────────────────────────────────
def generate_pdf(res: dict) -> bytes:
    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=A4)
    W, H = A4

    # Dark background
    c.setFillColorRGB(0.051, 0.051, 0.059)
    c.rect(0, 0, W, H, fill=1, stroke=0)

    # Gold bars
    c.setFillColorRGB(0.788, 0.659, 0.298)
    c.rect(0, H - 6, W, 6, fill=1, stroke=0)
    c.rect(0, 0, W, 4, fill=1, stroke=0)

    # Side lines
    c.setFillColorRGB(0.2, 0.165, 0.075)
    c.rect(28, 30, 1.5, H - 60, fill=1, stroke=0)
    c.rect(W - 30, 30, 1.5, H - 60, fill=1, stroke=0)

    # Title
    c.setFillColorRGB(0.941, 0.910, 0.847)
    c.setFont("Times-Italic", 46)
    c.drawCentredString(W / 2, H - 85, "AbiVegas 2026")

    c.setFont("Helvetica", 7)
    c.setFillColorRGB(0.541, 0.498, 0.431)
    c.drawCentredString(W / 2, H - 105, "K A R L - R E H B E I N - S C H U L E   H A N A U")

    # Divider
    c.setStrokeColorRGB(0.788, 0.659, 0.298)
    c.setLineWidth(0.5)
    c.line(55, H - 122, W - 55, H - 122)
    # Diamond
    c.setFillColorRGB(0.788, 0.659, 0.298)
    c.saveState()
    c.translate(W/2, H - 122)
    c.rotate(45)
    c.rect(-4, -4, 8, 8, fill=1, stroke=0)
    c.restoreState()

    # Subtitle
    c.setFont("Helvetica", 7)
    c.setFillColorRGB(0.541, 0.498, 0.431)
    c.drawCentredString(W / 2, H - 148, "R E S E R V I E R U N G S B E S T Ä T I G U N G")

    # Main info box
    bx, by, bw, bh = 55, H - 370, W - 110, 200
    c.setFillColorRGB(0.075, 0.075, 0.094)
    c.setStrokeColorRGB(0.788, 0.659, 0.298)
    c.setLineWidth(0.8)
    c.roundRect(bx, by, bw, bh, 4, fill=1, stroke=1)
    c.setFillColorRGB(0.788, 0.659, 0.298)
    c.rect(bx, by + bh - 3, bw, 3, fill=1, stroke=0)

    def row(label, value, y):
        c.setFont("Helvetica", 6.5)
        c.setFillColorRGB(0.541, 0.498, 0.431)
        c.drawString(bx + 20, y, label.upper())
        c.setFont("Times-Roman", 14)
        c.setFillColorRGB(0.941, 0.910, 0.847)
        c.drawString(bx + 20, y - 18, value)
        c.setStrokeColorRGB(0.15, 0.15, 0.2)
        c.setLineWidth(0.3)
        c.line(bx + 20, y - 24, bx + bw - 20, y - 24)

    ry = by + bh - 30
    row("Name",             res['name'],       ry)
    row("E-Mail-Adresse",   res['email'],      ry - 48)
    row("Tisch",            res['table_name'], ry - 96)
    row("Anzahl der Plaetze",
        str(res['seats']) + (" Platz" if res['seats'] == 1 else " Plaetze"), ry - 144)

    # Confirmation code
    cy = by - 75
    c.setFillColorRGB(0.075, 0.075, 0.094)
    c.setStrokeColorRGB(0.788, 0.659, 0.298)
    c.setLineWidth(0.7)
    c.roundRect(bx, cy, bw, 58, 4, fill=1, stroke=1)
    c.setFont("Helvetica", 6.5)
    c.setFillColorRGB(0.541, 0.498, 0.431)
    c.drawCentredString(W / 2, cy + 40, "B E S T Ä T I G U N G S C O D E")
    c.setFont("Courier-Bold", 20)
    c.setFillColorRGB(0.788, 0.659, 0.298)
    c.drawCentredString(W / 2, cy + 14, res['conf_code'])

    # Event details
    ey = cy - 65
    c.setFont("Helvetica", 6.5)
    c.setFillColorRGB(0.541, 0.498, 0.431)
    c.drawCentredString(W / 2, ey, "V E R A N S T A L T U N G S D E T A I L S")
    c.setStrokeColorRGB(0.2, 0.2, 0.25)
    c.setLineWidth(0.3)
    c.line(bx, ey - 8, bx + bw, ey - 8)

    details = [
        ("Datum",     "Freitag, 19. Juni 2026"),
        ("Uhrzeit",   "19:00 Uhr"),
        ("Ort",       "Congress Park Hanau"),
        ("Dresscode", "Overdressed"),
        ("Motto",     "AbiVegas"),
    ]
    dy = ey - 28
    for label, val in details:
        c.setFont("Helvetica-Bold", 7)
        c.setFillColorRGB(0.541, 0.498, 0.431)
        c.drawString(bx + 20, dy, label + ":")
        c.setFont("Helvetica", 9)
        c.setFillColorRGB(0.941, 0.910, 0.847)
        c.drawString(bx + 115, dy, val)
        dy -= 20

    # QR code
    try:
        qr = qrcode.QRCode(version=1, box_size=10, border=0)
        qr.add_data(res['conf_code'])
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")
        qr_buf = io.BytesIO()
        qr_img.save(qr_buf, 'PNG')
        qr_buf.seek(0)
        qr_size = 90
        qr_x = W - 140
        qr_y = cy - 130
        c.drawImage(ImageReader(qr_buf), qr_x, qr_y, width=qr_size, height=qr_size)
    except Exception as e:
        print(f"[QR CODE ERROR] {e}")
        pass

    # Footer
    c.setFont("Helvetica-Oblique", 7)
    c.setFillColorRGB(0.541, 0.498, 0.431)
    c.drawCentredString(W / 2, 60, f"Bitte diese Bestaetigung zum Einlass mitbringen  ·  Reserviert am {res['created_at']}")
    c.setFont("Helvetica", 6)
    c.setFillColorRGB(0.788, 0.659, 0.298)
    c.drawCentredString(W / 2, 45, "Created by Rayan Tajioui")

    # Watermark
    c.saveState()
    c.setFont("Times-Italic", 52)
    c.setFillColorRGB(0.788, 0.659, 0.298)
    c.setFillAlpha(0.06)
    c.translate(W/2, H/2)
    c.rotate(35)
    c.drawCentredString(0, 0, "AbiVegas 2026")
    c.restoreState()

    c.save()
    buf.seek(0)
    return buf.read()


# ── Helpers ───────────────────────────────────────────────────────────────────
def gen_code():
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))

def tables_with_availability():
    db = get_db()
    booked = {r['table_id']: r['total']
              for r in db.execute(
                  "SELECT table_id, SUM(seats) as total FROM reservations GROUP BY table_id")}
    return [{**t, 'booked': booked.get(t['id'], 0),
             'available': t['max_seats'] - booked.get(t['id'], 0)} for t in TABLES]

# ── Routes ────────────────────────────────────────────────────────────────────
@app.route('/')
def index(): return send_from_directory('static', 'index.html')

@app.route('/api/tables')
def api_tables(): return jsonify(tables_with_availability())

@app.route('/api/reserve', methods=['POST'])
def api_reserve():
    d        = request.get_json() or {}
    name     = (d.get('name') or '').strip()
    email    = (d.get('email') or '').strip().lower()
    table_id = d.get('table_id')
    seats    = int(d.get('seats', 1))

    if not all([name, email, table_id, seats]):
        return jsonify({'error': 'Fehlende Felder.'}), 400

    table = next((t for t in TABLES if t['id'] == table_id), None)
    if not table:
        return jsonify({'error': 'Ungueliger Tisch.'}), 400

    if table.get('teacher'):
        return jsonify({'error': 'Dieser Tisch ist nicht buchbar.'}), 409

    db = get_db()

    if db.execute("SELECT id FROM reservations WHERE LOWER(name)=LOWER(?)", (name,)).fetchone():
        return jsonify({'error': 'Dieser Name hat bereits eine Reservierung.'}), 409

    if db.execute("SELECT id FROM reservations WHERE email=?", (email,)).fetchone():
        return jsonify({'error': 'Diese E-Mail hat bereits eine Reservierung.'}), 409

    booked = db.execute(
        "SELECT COALESCE(SUM(seats),0) as t FROM reservations WHERE table_id=?",
        (table_id,)).fetchone()['t']

    if booked + seats > table['max_seats']:
        return jsonify({'error': f'Nur noch {table["max_seats"]-booked} Plaetze frei.'}), 409

    code = gen_code()
    while db.execute("SELECT id FROM reservations WHERE conf_code=?", (code,)).fetchone():
        code = gen_code()

    now = datetime.now().strftime('%d.%m.%Y %H:%M')
    db.execute("""INSERT INTO reservations
        (table_id,table_name,name,email,seats,created_at,conf_code)
        VALUES(?,?,?,?,?,?,?)""",
        (table_id, table['name'], name, email, seats, now, code))
    db.commit()

    res = {'name': name, 'email': email, 'table_id': table_id,
           'table_name': table['name'], 'seats': seats,
           'conf_code': code, 'created_at': now}

    pdf = generate_pdf(res)
    # send_email(res, pdf)  # Uncomment if email functionality is needed

    return jsonify({'success': True, 'conf_code': code, 'table_name': table['name']})

@app.route('/api/pdf/<code>')
def api_pdf(code):
    db  = get_db()
    row = db.execute("SELECT * FROM reservations WHERE conf_code=?", (code,)).fetchone()
    if not row: return jsonify({'error': 'Nicht gefunden'}), 404
    pdf = generate_pdf(dict(row))
    return send_file(io.BytesIO(pdf), mimetype='application/pdf',
                     download_name=f'Abiball2026_{code}.pdf', as_attachment=True)

# ── Admin ─────────────────────────────────────────────────────────────────────
def check_admin():
    return request.args.get('pw') == ADMIN_PW

@app.route('/api/admin/reservations')
def admin_list():
    if not check_admin(): return jsonify({'error': 'Unauth'}), 401
    rows = get_db().execute("SELECT * FROM reservations ORDER BY id DESC").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/admin/delete/<int:rid>', methods=['DELETE'])
def admin_delete(rid):
    if not check_admin(): return jsonify({'error': 'Unauth'}), 401
    get_db().execute("DELETE FROM reservations WHERE id=?", (rid,))
    get_db().commit()
    return jsonify({'success': True})

@app.route('/api/admin/export-excel')
def admin_export_excel():
    if not check_admin(): return jsonify({'error': 'Unauth'}), 401
    rows = get_db().execute("SELECT * FROM reservations ORDER BY created_at DESC").fetchall()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reservierungen"
    
    # Header
    headers = ["ID", "Tisch", "Name", "E-Mail", "Plätze", "Code", "Datum", "QR Code"]
    ws.append(headers)
    header_fill = PatternFill(start_color="B8972E", end_color="B8972E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Data rows
    temp_qr_dir = os.path.join(BASE_DIR, '.temp_qr')
    os.makedirs(temp_qr_dir, exist_ok=True)
    
    for idx, row in enumerate(rows, 2):
        ws.append([
            row['id'],
            row['table_name'],
            row['name'],
            row['email'],
            row['seats'],
            row['conf_code'],
            row['created_at'],
            ''
        ])
        
        # Generate QR code
        qr = qrcode.QRCode(version=1, box_size=5, border=1)
        qr.add_data(row['conf_code'])
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")
        qr_path = os.path.join(temp_qr_dir, f"qr_{row['id']}.png")
        qr_img.save(qr_path)
        
        # Insert QR code image
        try:
            img = XLImage(qr_path)
            img.width = 80
            img.height = 80
            ws.add_image(img, f"H{idx}")
            ws.row_dimensions[idx].height = 85
        except:
            pass
    
    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 20
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=f'Abiball_Reservierungen_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx', as_attachment=True)

# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    init_db()
    print("\n" + "="*55)
    print("  Abiball 2026 – Reservierungssystem")
    print("="*55)
    print("  URL:    http://localhost:5000")
    print("="*55 + "\n")
    app.run(host='0.0.0.0', debug=False, port=5000)